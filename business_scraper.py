"""
╔══════════════════════════════════════════════════════════════════════════════╗
║         NEURAL AUTOMATER  //  INTELLIGENT BUSINESS SCRAPER ENGINE           ║
║      Google Maps · Google Business · LinkedIn · Instagram → Excel/CSV       ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import asyncio
import pandas as pd
import json
import re
import time
import random
import os
from datetime import datetime
from pathlib import Path


class BusinessScraper:
    """
    World-class intelligent business scraper.
    Sources: Google Maps, LinkedIn Companies, Instagram Business profiles.
    Output: Pandas DataFrame → Excel (.xlsx) or CSV.
    """

    # Columns exported in the sheet
    COLUMNS = [
        "Name", "Category", "Address", "Phone", "Website",
        "Rating", "Reviews", "Hours", "Email", "Source",
        "LinkedIn", "Instagram", "Followers", "Description", "Scraped At"
    ]

    def __init__(self, browser_agent, llm_client=None, log_callback=None):
        """
        :param browser_agent: The shared BrowserAgent instance from the app.
        :param llm_client: Optional LLMClient for AI enrichment.
        :param log_callback: Optional callable(str) for live progress logging to GUI.
        """
        self.agent = browser_agent
        self.llm = llm_client
        self.log = log_callback or print
        self._records = []
        self._stop_requested = False

    def stop(self):
        """Request the scraper to stop after the current record."""
        self._stop_requested = True

    def _new_record(self, **kwargs):
        rec = {col: "" for col in self.COLUMNS}
        rec["Scraped At"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        rec.update(kwargs)
        return rec

    def _human_delay(self, min_s=0.8, max_s=2.2):
        """Random human-like sleep to avoid bot detection."""
        time.sleep(random.uniform(min_s, max_s))

    # ═══════════════════════════════════════════════════════════════════════════
    #  GOOGLE MAPS SCRAPER
    # ═══════════════════════════════════════════════════════════════════════════

    async def scrape_google_maps(self, query: str, max_results: int = 30) -> list:
        """
        Scrapes businesses from Google Maps for a given search query.
        Returns a list of business record dicts.
        """
        self._stop_requested = False
        records = []
        self.log(f"[MAPS] Searching Google Maps: '{query}' (max={max_results})")

        try:
            page = self.agent.page
            if not page:
                self.log("[MAPS] ERROR: Browser not started. Launch browser first.")
                return records

            search_url = f"https://www.google.com/maps/search/{query.replace(' ', '+')}"
            await page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
            await asyncio.sleep(2)

            # Dismiss cookie consent if present
            try:
                await page.click("button[aria-label='Accept all']", timeout=3000)
                await asyncio.sleep(1)
            except Exception:
                pass

            self.log("[MAPS] Page loaded. Scrolling results list...")

            # The sidebar panel selector for Google Maps results
            panel_sel = "div[role='feed']"
            await page.wait_for_selector(panel_sel, timeout=15000)

            seen_names = set()
            scroll_attempts = 0
            max_scroll_attempts = 50

            while len(records) < max_results and scroll_attempts < max_scroll_attempts:
                if self._stop_requested:
                    self.log("[MAPS] Stop requested by user.")
                    break

                # Grab all result items currently visible
                items = await page.query_selector_all(
                    "div[role='feed'] > div > div > a[href*='/maps/place/']"
                )

                new_found = False
                for item in items:
                    if len(records) >= max_results or self._stop_requested:
                        break

                    try:
                        name = await item.get_attribute("aria-label") or ""
                        href = await item.get_attribute("href") or ""

                        if not name or name in seen_names:
                            continue

                        seen_names.add(name)
                        new_found = True

                        self.log(f"[MAPS] ({len(records)+1}/{max_results}) Extracting: {name}")

                        # Click to open the detail panel
                        await item.click()
                        await asyncio.sleep(1.5)

                        rec = await self._extract_maps_detail(page, name)
                        rec["Source"] = "Google Maps"
                        records.append(rec)

                    except Exception as e:
                        self.log(f"[MAPS] Item extract error: {e}")
                        continue

                if not new_found or len(records) >= max_results:
                    break

                # Scroll the feed panel to load more results
                await page.evaluate(f"""
                    const feed = document.querySelector("div[role='feed']");
                    if (feed) feed.scrollTop += 800;
                """)
                self._human_delay(1.5, 3.0)
                scroll_attempts += 1

        except Exception as e:
            self.log(f"[MAPS] Fatal error: {e}")

        self.log(f"[MAPS] Scrape complete. {len(records)} businesses found.")
        return records

    async def _extract_maps_detail(self, page, fallback_name: str) -> dict:
        """Extract detailed info from the currently open Maps detail panel."""
        rec = self._new_record(Name=fallback_name)
        try:
            await asyncio.sleep(0.8)
            # Name
            try:
                name_el = await page.query_selector("h1.DUwDvf")
                if name_el:
                    rec["Name"] = (await name_el.inner_text()).strip()
            except Exception:
                pass

            # Category
            try:
                cat_els = await page.query_selector_all("button.DkEaL")
                if cat_els:
                    cats = [await el.inner_text() for el in cat_els]
                    rec["Category"] = ", ".join(cats).strip()
            except Exception:
                pass

            # Address
            try:
                addr_el = await page.query_selector("button[data-item-id='address']")
                if addr_el:
                    rec["Address"] = (await addr_el.get_attribute("aria-label") or "").replace("Address: ", "").strip()
            except Exception:
                pass

            # Phone
            try:
                # Primary: Map's robust data-item-id
                phone_el = await page.query_selector("button[data-item-id^='phone:']")
                if not phone_el:
                    # Fallback: Tooltip
                    phone_el = await page.query_selector("button[data-tooltip*='phone' i]")
                    
                if phone_el:
                    # Try to extract the clean tel: link first
                    item_id = await phone_el.get_attribute("data-item-id") or ""
                    if "tel:" in item_id:
                        rec["Phone"] = item_id.split("tel:")[-1].strip()
                    else:
                        # Fallback to visible text
                        text = (await phone_el.inner_text()).strip()
                        if "\n" in text:
                            rec["Phone"] = text.split("\n")[0].strip()
                        elif text:
                            rec["Phone"] = text
                        else:
                            # Final fallback: aria-label
                            lbl = await phone_el.get_attribute("aria-label") or ""
                            rec["Phone"] = re.sub(r"(Phone:|Call)\s*", "", lbl, flags=re.IGNORECASE).strip()
            except Exception:
                pass

            # Website
            try:
                web_el = await page.query_selector("a[data-item-id='authority']")
                if web_el:
                    rec["Website"] = (await web_el.get_attribute("href") or "").strip()
            except Exception:
                pass

            # Rating
            try:
                rating_el = await page.query_selector("div.F7nice span[aria-hidden='true']")
                if rating_el:
                    rec["Rating"] = (await rating_el.inner_text()).strip()
            except Exception:
                pass

            # Reviews count
            try:
                rev_el = await page.query_selector("div.F7nice span[aria-label*='review']")
                if rev_el:
                    rev_text = (await rev_el.get_attribute("aria-label") or "")
                    rec["Reviews"] = re.sub(r"[^0-9,]", "", rev_text).strip()
            except Exception:
                pass

            # Hours
            try:
                hrs_el = await page.query_selector("div[data-hide-tooltip-on-mouse-move] .t39EBf")
                if hrs_el:
                    rec["Hours"] = (await hrs_el.inner_text()).strip()
            except Exception:
                pass

        except Exception as e:
            self.log(f"[MAPS] Detail extraction error: {e}")

        return rec

    # ═══════════════════════════════════════════════════════════════════════════
    #  LINKEDIN COMPANY SCRAPER
    # ═══════════════════════════════════════════════════════════════════════════

    async def scrape_linkedin_companies(self, query: str, max_results: int = 20) -> list:
        """
        Scrapes LinkedIn company search results.
        Requires being logged into LinkedIn via the browser.
        """
        self._stop_requested = False
        records = []
        self.log(f"[LINKEDIN] Searching LinkedIn companies: '{query}' (max={max_results})")

        try:
            page = self.agent.page
            if not page:
                self.log("[LINKEDIN] ERROR: Browser not started.")
                return records

            search_url = (
                f"https://www.linkedin.com/search/results/companies/"
                f"?keywords={query.replace(' ', '%20')}"
            )
            await page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
            await asyncio.sleep(3)

            # Check if we're redirected to login
            current_url = page.url
            if "login" in current_url or "authwall" in current_url:
                self.log("[LINKEDIN] Not logged in — redirected to login page. Please log in via the ACCOUNTS tab first.")
                return records

            page_num = 1
            while len(records) < max_results:
                if self._stop_requested:
                    break

                self.log(f"[LINKEDIN] Processing page {page_num}...")

                # Wait for results container
                try:
                    await page.wait_for_selector("ul.reusable-search__entity-result-list", timeout=10000)
                except Exception:
                    self.log("[LINKEDIN] Could not find results list. Possibly blocked or no results.")
                    break

                items = await page.query_selector_all(
                    "li.reusable-search__result-container"
                )

                if not items:
                    break

                for item in items:
                    if len(records) >= max_results or self._stop_requested:
                        break

                    try:
                        rec = self._new_record()
                        rec["Source"] = "LinkedIn"

                        # Name
                        try:
                            name_el = await item.query_selector("span.entity-result__title-text a span[aria-hidden='true']")
                            if name_el:
                                rec["Name"] = (await name_el.inner_text()).strip()
                        except Exception:
                            pass

                        # LinkedIn profile URL
                        try:
                            link_el = await item.query_selector("a.app-aware-link")
                            if link_el:
                                rec["LinkedIn"] = (await link_el.get_attribute("href") or "").split("?")[0]
                        except Exception:
                            pass

                        # Category / industry + location snippet
                        try:
                            subtitle_el = await item.query_selector("div.entity-result__primary-subtitle")
                            if subtitle_el:
                                rec["Category"] = (await subtitle_el.inner_text()).strip()
                        except Exception:
                            pass

                        try:
                            secondary_el = await item.query_selector("div.entity-result__secondary-subtitle")
                            if secondary_el:
                                rec["Address"] = (await secondary_el.inner_text()).strip()
                        except Exception:
                            pass

                        # Description / summary
                        try:
                            desc_el = await item.query_selector("p.entity-result__summary")
                            if desc_el:
                                rec["Description"] = (await desc_el.inner_text()).strip()
                        except Exception:
                            pass

                        if rec["Name"]:
                            records.append(rec)
                            self.log(f"[LINKEDIN] ({len(records)}/{max_results}) {rec['Name']}")

                    except Exception as e:
                        self.log(f"[LINKEDIN] Item error: {e}")

                # Try next page
                try:
                    next_btn = await page.query_selector("button[aria-label='Next']")
                    if next_btn and len(records) < max_results:
                        await next_btn.click()
                        page_num += 1
                        self._human_delay(2.5, 4.0)
                    else:
                        break
                except Exception:
                    break

        except Exception as e:
            self.log(f"[LINKEDIN] Fatal error: {e}")

        self.log(f"[LINKEDIN] Scrape complete. {len(records)} companies found.")
        return records

    # ═══════════════════════════════════════════════════════════════════════════
    #  INSTAGRAM BUSINESS SCRAPER
    # ═══════════════════════════════════════════════════════════════════════════

    async def scrape_instagram_profiles(self, query: str, max_results: int = 20) -> list:
        """
        Searches Instagram via the Explore / Hashtag or account search.
        Extracts business profile info.
        Requires being logged into Instagram via the browser.
        """
        self._stop_requested = False
        records = []
        self.log(f"[INSTAGRAM] Searching Instagram: '{query}' (max={max_results})")

        try:
            page = self.agent.page
            if not page:
                self.log("[INSTAGRAM] ERROR: Browser not started.")
                return records

            # Use the search page
            search_url = f"https://www.instagram.com/explore/search/keyword/?q={query.replace(' ', '%20')}"
            await page.goto(search_url, wait_until="domcontentloaded", timeout=30000)
            await asyncio.sleep(3)

            current_url = page.url
            if "login" in current_url or "accounts/login" in current_url:
                self.log("[INSTAGRAM] Not logged in. Please log in via the ACCOUNTS tab first.")
                return records

            self.log("[INSTAGRAM] Collecting profile links from search results...")

            # Collect account links from search
            profile_links = set()

            # Try the accounts tab in search
            try:
                accounts_tab = await page.query_selector("span:text('Accounts')")
                if accounts_tab:
                    await accounts_tab.click()
                    await asyncio.sleep(2)
            except Exception:
                pass

            scroll_count = 0
            while len(profile_links) < max_results * 2 and scroll_count < 10:
                links = await page.query_selector_all("a[href^='/'][href$='/']")
                for link in links:
                    href = await link.get_attribute("href")
                    if href and href.count("/") == 2 and href != "/" and "explore" not in href:
                        profile_links.add(f"https://www.instagram.com{href}")
                if len(profile_links) >= max_results * 2:
                    break
                await page.evaluate("window.scrollBy(0, 600)")
                await asyncio.sleep(1.5)
                scroll_count += 1

            self.log(f"[INSTAGRAM] Found {len(profile_links)} profile links. Visiting each...")

            for link in list(profile_links)[:max_results]:
                if self._stop_requested:
                    break

                try:
                    await page.goto(link, wait_until="domcontentloaded", timeout=20000)
                    await asyncio.sleep(2)

                    rec = self._new_record()
                    rec["Source"] = "Instagram"
                    rec["Instagram"] = link

                    # Username
                    try:
                        username = link.rstrip("/").split("/")[-1]
                        rec["Name"] = f"@{username}"
                    except Exception:
                        pass

                    # Bio / description
                    try:
                        bio_el = await page.query_selector("div._ap3a._aaco._aacu._aacx._aad7._aade span")
                        if bio_el:
                            rec["Description"] = (await bio_el.inner_text()).strip()
                    except Exception:
                        pass

                    # External website link in bio
                    try:
                        web_els = await page.query_selector_all("a[href^='https://l.instagram.com']")
                        if web_els:
                            raw = await web_els[0].get_attribute("href") or ""
                            # Decode the redirect URL
                            match = re.search(r"u=([^&]+)", raw)
                            if match:
                                from urllib.parse import unquote
                                rec["Website"] = unquote(match.group(1))
                    except Exception:
                        pass

                    # Followers count
                    try:
                        meta = await page.query_selector_all("meta[name='description']")
                        if meta:
                            content = await meta[0].get_attribute("content") or ""
                            # content format: "123K Followers, 456 Following, 789 Posts - ..."
                            match = re.search(r"([\d,\.]+[KMB]?)\s+Followers", content, re.IGNORECASE)
                            if match:
                                rec["Followers"] = match.group(1)
                    except Exception:
                        pass

                    # Category (shown under username for business profiles)
                    try:
                        cat_el = await page.query_selector("div._ap3a._aaco._aacu._aacx._aad7._aade:not(:has(a)):not(:has(button))")
                        if cat_el:
                            text = (await cat_el.inner_text()).strip()
                            if text and len(text) < 60:
                                rec["Category"] = text
                    except Exception:
                        pass

                    records.append(rec)
                    self.log(f"[INSTAGRAM] ({len(records)}/{max_results}) {rec['Name']} | Followers: {rec['Followers']}")
                    self._human_delay(1.5, 3.0)

                except Exception as e:
                    self.log(f"[INSTAGRAM] Profile error for {link}: {e}")
                    continue

        except Exception as e:
            self.log(f"[INSTAGRAM] Fatal error: {e}")

        self.log(f"[INSTAGRAM] Scrape complete. {len(records)} profiles found.")
        return records

    # ═══════════════════════════════════════════════════════════════════════════
    #  AI ENRICHMENT
    # ═══════════════════════════════════════════════════════════════════════════

    def enrich_with_ai(self, records: list) -> list:
        """
        Use the configured LLM to fill in missing fields intelligently.
        Runs synchronously — call from a background thread.
        """
        if not self.llm:
            self.log("[AI] No LLM configured — skipping enrichment.")
            return records

        self.log(f"[AI] Enriching {len(records)} records with AI...")

        for i, rec in enumerate(records):
            if self._stop_requested:
                break

            # Only enrich if there's something to work with
            if not rec.get("Name"):
                continue

            missing = [k for k in ["Category", "Description", "Email"] if not rec.get(k)]
            if not missing:
                continue

            prompt = (
                "You are a business intelligence analyst. Given the following scraped business data, "
                f"fill in any MISSING fields using best-guess inference. Return ONLY a JSON object with these keys: {missing}.\n\n"
                f"Business data:\n{json.dumps({k: v for k, v in rec.items() if v}, indent=2)}\n\n"
                "Return only valid JSON with the missing field values. No explanation, no markdown."
            )

            try:
                response = self.llm.generate_text(prompt)
                # Extract JSON from response
                match = re.search(r"\{[^}]+\}", response, re.DOTALL)
                if match:
                    filled = json.loads(match.group())
                    for key, val in filled.items():
                        if key in rec and not rec[key]:
                            rec[key] = str(val).strip()
                self.log(f"[AI] Enriched ({i+1}/{len(records)}): {rec['Name']}")
            except Exception as e:
                self.log(f"[AI] Enrichment error for record {i}: {e}")

        self.log("[AI] Enrichment complete.")
        return records

    # ═══════════════════════════════════════════════════════════════════════════
    #  EXPORT METHODS
    # ═══════════════════════════════════════════════════════════════════════════

    def export_to_excel(self, records: list, filepath: str) -> bool:
        """
        Export scraped records to a beautifully formatted Excel file.
        Uses openpyxl for rich formatting.
        """
        try:
            import openpyxl
            from openpyxl.styles import (
                Font, PatternFill, Alignment, Border, Side
            )
            from openpyxl.utils import get_column_letter

            df = pd.DataFrame(records, columns=self.COLUMNS)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Business Intelligence"

            # ── Color Palette ─────────────────────────
            HEADER_BG   = "050508"   # Void black
            HEADER_FG   = "00f0ff"   # Electric cyan
            ROW_ALT_BG  = "0e0e1a"   # Dark navy
            ROW_BG      = "131325"   # Slightly lighter
            TEXT_COLOR  = "e8eaf6"   # Soft white
            BORDER_CLR  = "1e1e38"   # Subtle border
            GOLD        = "ffc857"   # Gold metallic

            header_fill   = PatternFill("solid", fgColor=HEADER_BG)
            row_fill_a    = PatternFill("solid", fgColor=ROW_ALT_BG)
            row_fill_b    = PatternFill("solid", fgColor=ROW_BG)
            thin_border   = Border(
                left=Side(style="thin", color=BORDER_CLR),
                right=Side(style="thin", color=BORDER_CLR),
                top=Side(style="thin", color=BORDER_CLR),
                bottom=Side(style="thin", color=BORDER_CLR),
            )

            # ── Title Row ─────────────────────────────
            ws.merge_cells(f"A1:{get_column_letter(len(self.COLUMNS))}1")
            title_cell = ws["A1"]
            title_cell.value = (
                f"⚡ NEURAL AUTOMATER  //  BUSINESS INTELLIGENCE REPORT  "
                f"[{datetime.now().strftime('%Y-%m-%d %H:%M')}]  "
                f"//  {len(records)} RECORDS"
            )
            title_cell.font = Font(
                name="Consolas", size=13, bold=True, color=GOLD
            )
            title_cell.fill = PatternFill("solid", fgColor=HEADER_BG)
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[1].height = 24

            # ── Header Row ────────────────────────────
            for col_idx, col_name in enumerate(self.COLUMNS, start=1):
                cell = ws.cell(row=2, column=col_idx, value=col_name)
                cell.font = Font(name="Consolas", size=11, bold=True, color=HEADER_FG)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            ws.row_dimensions[2].height = 22

            # ── Data Rows ─────────────────────────────
            for row_idx, record in enumerate(records, start=3):
                fill = row_fill_a if row_idx % 2 == 0 else row_fill_b
                for col_idx, col_name in enumerate(self.COLUMNS, start=1):
                    val = record.get(col_name, "")
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.font = Font(name="Consolas", size=10, color=TEXT_COLOR)
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                    cell.border = thin_border
                ws.row_dimensions[row_idx].height = 18

            # ── Auto-size Columns ─────────────────────
            col_widths = {
                "Name": 28, "Category": 22, "Address": 35, "Phone": 18,
                "Website": 35, "Rating": 9, "Reviews": 10, "Hours": 20,
                "Email": 28, "Source": 14, "LinkedIn": 40, "Instagram": 32,
                "Followers": 12, "Description": 45, "Scraped At": 20,
            }
            for col_idx, col_name in enumerate(self.COLUMNS, start=1):
                width = col_widths.get(col_name, 18)
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            # Freeze header rows
            ws.freeze_panes = "A3"

            # Auto-filter
            ws.auto_filter.ref = (
                f"A2:{get_column_letter(len(self.COLUMNS))}{len(records)+2}"
            )

            wb.save(filepath)
            self.log(f"[EXPORT] Excel saved → {filepath}")
            return True

        except ImportError:
            self.log("[EXPORT] ERROR: openpyxl not installed. Run: pip install openpyxl")
            return False
        except Exception as e:
            self.log(f"[EXPORT] Excel export error: {e}")
            return False

    def export_to_csv(self, records: list, filepath: str) -> bool:
        """Export records to a clean CSV file."""
        try:
            df = pd.DataFrame(records, columns=self.COLUMNS)
            df.to_csv(filepath, index=False, encoding="utf-8-sig")
            self.log(f"[EXPORT] CSV saved → {filepath}")
            return True
        except Exception as e:
            self.log(f"[EXPORT] CSV export error: {e}")
            return False

    def get_records(self) -> list:
        return self._records
